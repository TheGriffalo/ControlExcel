import openpyxl

from openpyxl import Workbook

#Create the workbook object
wb = Workbook()

#Create a new sheet
ws = wb.create_sheet("A Sheet", 0) #Insert at first position

#Change the name of the sheet
ws.title = "Hello World"

#Create another sheet
ws2 = wb.create_sheet("Sheet nr 2")



wb.remove(wb["Sheet"])


wb.save("Lesson_9.xlsx")

for sheet in wb:
    print(sheet.title)